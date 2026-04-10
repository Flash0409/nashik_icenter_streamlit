import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from io import BytesIO
import json
import os
import re
from datetime import datetime, timedelta
import warnings

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────
# CONFIG & PAGE SETUP
# ─────────────────────────────────────────────────
st.set_page_config(
    page_title="Material Shortage Forecaster",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

SAVED_LOTS_FILE = "saved_lots.json"

# ─────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

    :root {
        --bg-dark: #0a0e17;
        --bg-card: #111827;
        --bg-card-hover: #1a2235;
        --accent-blue: #3b82f6;
        --accent-cyan: #06b6d4;
        --accent-emerald: #10b981;
        --accent-amber: #f59e0b;
        --accent-red: #ef4444;
        --accent-purple: #8b5cf6;
        --text-primary: #f1f5f9;
        --text-secondary: #94a3b8;
        --border-color: #1e293b;
    }

    .stApp {
        font-family: 'DM Sans', sans-serif;
    }

    .main-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 50%, #0f172a 100%);
        border: 1px solid #2d2b55;
        border-radius: 16px;
        padding: 2rem 2.5rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: radial-gradient(ellipse at 20% 50%, rgba(59,130,246,0.08) 0%, transparent 60%),
                    radial-gradient(ellipse at 80% 50%, rgba(139,92,246,0.08) 0%, transparent 60%);
        pointer-events: none;
    }
    .main-header h1 {
        margin: 0; font-size: 1.8rem; font-weight: 700;
        background: linear-gradient(90deg, #e2e8f0, #93c5fd);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .main-header p {
        margin: 0.3rem 0 0 0; color: #94a3b8; font-size: 0.95rem;
    }

    .metric-card {
        background: linear-gradient(145deg, #111827, #1a2235);
        border: 1px solid #1e293b;
        border-radius: 12px;
        padding: 1.25rem 1.5rem;
        text-align: center;
        transition: all 0.2s ease;
    }
    .metric-card:hover { border-color: #3b82f6; transform: translateY(-2px); }
    .metric-value {
        font-size: 2rem; font-weight: 700;
        font-family: 'JetBrains Mono', monospace;
    }
    .metric-label {
        font-size: 0.8rem; color: #94a3b8; text-transform: uppercase;
        letter-spacing: 0.05em; margin-top: 0.25rem;
    }
    .metric-blue .metric-value { color: #3b82f6; }
    .metric-emerald .metric-value { color: #10b981; }
    .metric-amber .metric-value { color: #f59e0b; }
    .metric-red .metric-value { color: #ef4444; }
    .metric-purple .metric-value { color: #8b5cf6; }
    .metric-cyan .metric-value { color: #06b6d4; }

    .section-header {
        font-size: 1.15rem; font-weight: 600; color: #e2e8f0;
        border-bottom: 2px solid #3b82f6; padding-bottom: 0.5rem;
        margin: 1.5rem 0 1rem 0;
    }

    .lot-chip {
        display: inline-block;
        background: rgba(139,92,246,0.15);
        color: #c4b5fd;
        border: 1px solid rgba(139,92,246,0.3);
        border-radius: 6px;
        padding: 0.15rem 0.5rem;
        font-size: 0.72rem;
        font-weight: 600;
        margin: 1px 2px;
    }

    div[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f172a 0%, #1e1b4b 100%);
    }

    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────
def load_saved_lots() -> dict:
    if os.path.exists(SAVED_LOTS_FILE):
        with open(SAVED_LOTS_FILE, "r") as f:
            raw_lots = json.load(f)

        normalized_lots = {}
        changed = False
        for lot_name, lot_info in raw_lots.items():
            if not isinstance(lot_info, dict):
                changed = True
                continue
            project_name = str(lot_info.get("project_name", lot_info.get("project", ""))).strip()
            row_number = str(lot_info.get("row_number", lot_info.get("rom_number", ""))).strip()
            normalized_lots[str(lot_name).strip()] = {
                "project_name": project_name,
                "row_number": row_number,
            }
            if lot_info != normalized_lots[str(lot_name).strip()]:
                changed = True

        if changed:
            save_lots(normalized_lots)
        return normalized_lots
    return {}


def save_lots(lots: dict):
    with open(SAVED_LOTS_FILE, "w") as f:
        json.dump(lots, f, indent=2, default=str)


def get_lot_project_name(lot_info: dict) -> str:
    if not isinstance(lot_info, dict):
        return ""
    return str(lot_info.get("project_name", lot_info.get("project", ""))).strip()


def get_lot_row_number(lot_info: dict) -> str:
    if not isinstance(lot_info, dict):
        return ""
    return str(lot_info.get("row_number", lot_info.get("rom_number", ""))).strip()


def get_lot_bom_scope(df_bom_input: pd.DataFrame, lot_info: dict, proj_name_col: str | None, proj_num_col: str | None) -> pd.DataFrame:
    lot_bom_project = get_lot_project_name(lot_info)
    df_lot_bom = df_bom_input.copy()
    if lot_bom_project and lot_bom_project != "-- All --":
        p_col = proj_name_col or proj_num_col
        if p_col and p_col in df_lot_bom.columns:
            df_lot_bom = df_lot_bom[df_lot_bom[p_col].astype(str).str.strip() == lot_bom_project]
    return df_lot_bom


def compute_lot_metrics(df_bom_input: pd.DataFrame, lot_info: dict) -> dict:
    scoped_bom = get_lot_bom_scope(df_bom_input, lot_info, bom_proj_name_col, bom_proj_num_col)
    if scoped_bom.empty:
        return {"component_count": 0, "short_components": 0, "total_open_qty": 0}
    if not bom_comp_code_col or not bom_open_qty_col or bom_comp_code_col not in scoped_bom.columns or bom_open_qty_col not in scoped_bom.columns:
        return {"component_count": 0, "short_components": 0, "total_open_qty": 0}

    grp = [bom_comp_code_col]
    if bom_comp_desc_col and bom_comp_desc_col in scoped_bom.columns:
        grp.append(bom_comp_desc_col)
    lot_comp = scoped_bom.groupby(grp, dropna=False).agg(
        Open_Qty=(bom_open_qty_col, "sum"),
        Total_Available=("_total_available", "max"),
    ).reset_index()
    lot_comp["Gap"] = lot_comp["Total_Available"] - lot_comp["Open_Qty"]

    return {
        "component_count": int(len(lot_comp)),
        "short_components": int((lot_comp["Gap"] < 0).sum()),
        "total_open_qty": int(lot_comp["Open_Qty"].sum()),
    }


def _first_non_empty(series: pd.Series):
    non_empty = series.dropna().astype(str).map(str.strip)
    non_empty = non_empty[non_empty != ""]
    if non_empty.empty:
        return None
    return non_empty.iloc[0]


def build_po_overview_dataframe(shortage_df_input: pd.DataFrame, df_bom_input: pd.DataFrame) -> pd.DataFrame:
    if shortage_df_input.empty or not bom_comp_code_col:
        return pd.DataFrame()

    bom_scope = df_bom_input.copy()
    bom_scope[bom_comp_code_col] = bom_scope[bom_comp_code_col].astype(str).str.strip()
    bom_project_col = bom_proj_name_col or bom_proj_num_col
    if bom_project_col and bom_project_col in bom_scope.columns:
        bom_scope[bom_project_col] = bom_scope[bom_project_col].astype(str).str.strip()

    for id_col in [bom_order_col, bom_wo_col, bom_cust_po_col]:
        if id_col and id_col in bom_scope.columns:
            bom_scope[id_col] = normalize_identifier_series(bom_scope[id_col])

    po_rows = shortage_df_input.copy()
    po_rows["Project"] = po_rows["Project"].astype(str).str.strip()
    po_rows["Component Code"] = po_rows["Component Code"].astype(str).str.strip()

    bom_match_rows = bom_scope.copy()
    if bom_project_col and bom_project_col in bom_match_rows.columns and bom_project_col in po_rows.columns:
        bom_match_rows[bom_project_col] = bom_match_rows[bom_project_col].astype(str).str.strip()

    merge_left_keys = ["Component Code"]
    merge_right_keys = [bom_comp_code_col]
    if bom_project_col and bom_project_col in bom_match_rows.columns and bom_project_col in po_rows.columns:
        merge_left_keys = ["Project", "Component Code"]
        merge_right_keys = [bom_project_col, bom_comp_code_col]

    merged = po_rows.merge(
        bom_match_rows,
        left_on=merge_left_keys,
        right_on=merge_right_keys,
        how="left",
        suffixes=("", "_bom"),
    )

    if bom_incoming_po_col and bom_incoming_po_col in merged.columns:
        merged["Incoming PO Quantity"] = pd.to_numeric(merged[bom_incoming_po_col], errors="coerce").fillna(0)
    else:
        merged["Incoming PO Quantity"] = 0

    if bom_po_recv_col and bom_po_recv_col in merged.columns:
        merged["PO in Receiving"] = pd.to_numeric(merged[bom_po_recv_col], errors="coerce").fillna(0)
    else:
        merged["PO in Receiving"] = 0

    if bom_po_promise_col and bom_po_promise_col in merged.columns:
        merged["PO Promise Date"] = pd.to_datetime(merged[bom_po_promise_col], errors="coerce")
    else:
        merged["PO Promise Date"] = pd.NaT

    if bom_buyer_col and bom_buyer_col in merged.columns:
        merged["Buyer Name"] = merged[bom_buyer_col].astype(str).str.strip().replace({"nan": "", "None": ""})
    else:
        merged["Buyer Name"] = ""

    if bom_po_generated_col and bom_po_generated_col in merged.columns:
        merged["PO Generated Raw"] = merged[bom_po_generated_col].astype(str).str.strip().replace({"nan": "", "None": ""})
    else:
        merged["PO Generated Raw"] = ""

    if bom_wo_col and bom_wo_col in merged.columns:
        merged.rename(columns={bom_wo_col: "Work Order Number"}, inplace=True)
    else:
        merged["Work Order Number"] = ""

    if bom_order_col and bom_order_col in merged.columns:
        merged.rename(columns={bom_order_col: "Order Number"}, inplace=True)
    else:
        merged["Order Number"] = ""

    if bom_cust_po_col and bom_cust_po_col in merged.columns:
        merged.rename(columns={bom_cust_po_col: "Customer PO"}, inplace=True)
    else:
        merged["Customer PO"] = ""

    merged["PO Generated"] = merged.apply(
        lambda row: "Yes" if (
            str(row.get("PO Generated Raw", "")).strip() not in {"", "nan", "None"}
            or pd.to_numeric(row.get("Incoming PO Quantity"), errors="coerce") > 0
            or pd.to_numeric(row.get("PO in Receiving"), errors="coerce") > 0
            or pd.notna(row.get("PO Promise Date"))
        ) else "No",
        axis=1,
    )

    def join_unique(values) -> str:
        cleaned = []
        for value in values:
            if pd.isna(value):
                continue
            text = str(value).strip()
            if text in {"", "nan", "None"}:
                continue
            if text not in cleaned:
                cleaned.append(text)
        return ", ".join(cleaned)

    club_keys = ["Project", "Lot", "PO Generated", "Buyer Name", "Incoming PO Quantity", "PO in Receiving", "PO Promise Date"]
    club_keys = [column for column in club_keys if column in merged.columns]

    club_agg = {}
    if "Shortage" in merged.columns:
        club_agg["Shortage"] = "sum"
    if "Open Qty" in merged.columns:
        club_agg["Open Qty"] = "sum"
    if "Component Code" in merged.columns:
        club_agg["Component Code"] = join_unique
    if "Component Desc" in merged.columns:
        club_agg["Component Desc"] = join_unique
    if "Work Order Number" in merged.columns:
        club_agg["Work Order Number"] = join_unique
    if "Order Number" in merged.columns:
        club_agg["Order Number"] = join_unique
    if "Customer PO" in merged.columns:
        club_agg["Customer PO"] = join_unique

    if not club_keys or not club_agg:
        display_cols = [
            "Project", "Lot", "Component Code", "Component Desc", "Shortage", "Open Qty",
            "Work Order Number", "Order Number", "Customer PO",
            "PO Generated", "PO Promise Date", "Buyer Name", "Incoming PO Quantity", "PO in Receiving",
        ]
        existing_display_cols = [c for c in display_cols if c in merged.columns]
        merged = merged[existing_display_cols].copy()
        merged["PO Promise Date"] = pd.to_datetime(merged["PO Promise Date"], errors="coerce")
        merged["Incoming PO Quantity"] = pd.to_numeric(merged["Incoming PO Quantity"], errors="coerce").fillna(0)
        merged["PO in Receiving"] = pd.to_numeric(merged["PO in Receiving"], errors="coerce").fillna(0)
        return merged

    clubbed = merged.groupby(club_keys, dropna=False, as_index=False).agg(club_agg)
    clubbed["Clubbed Rows"] = merged.groupby(club_keys, dropna=False).size().values

    display_cols = [
        "Project", "Lot", "PO Generated", "Buyer Name", "Incoming PO Quantity", "PO in Receiving",
        "PO Promise Date", "Clubbed Rows", "Work Order Number", "Order Number", "Customer PO",
        "Component Code", "Component Desc", "Shortage", "Open Qty",
    ]
    existing_display_cols = [c for c in display_cols if c in clubbed.columns]
    clubbed = clubbed[existing_display_cols].copy()
    clubbed["PO Promise Date"] = pd.to_datetime(clubbed["PO Promise Date"], errors="coerce")
    clubbed["Incoming PO Quantity"] = pd.to_numeric(clubbed["Incoming PO Quantity"], errors="coerce").fillna(0)
    clubbed["PO in Receiving"] = pd.to_numeric(clubbed["PO in Receiving"], errors="coerce").fillna(0)
    return clubbed


def robust_to_datetime(series: pd.Series) -> pd.Series:
    if series.dropna().empty:
        return pd.to_datetime(series, errors="coerce")

    sample = series.dropna().head(50)
    best_result_sample = None
    best_count = -1
    best_kwargs = {}

    strategies = [
        {"dayfirst": False},
        {"dayfirst": True},
    ]
    for kwargs in strategies:
        result = pd.to_datetime(sample, errors="coerce", **kwargs)
        count = result.notna().sum()
        if count > best_count:
            best_count = count
            best_kwargs = kwargs
            best_result_sample = result

    best_mixed_kwargs = None
    for dayfirst in [False, True]:
        try:
            result = pd.to_datetime(sample, errors="coerce", format="mixed", dayfirst=dayfirst)
            count = result.notna().sum()
            if count > best_count:
                best_count = count
                best_mixed_kwargs = {"format": "mixed", "dayfirst": dayfirst}
        except Exception:
            pass

    if best_mixed_kwargs:
        try:
            return pd.to_datetime(series, errors="coerce", **best_mixed_kwargs)
        except Exception:
            pass
    return pd.to_datetime(series, errors="coerce", **best_kwargs)


def parse_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def normalize_identifier_series(series: pd.Series) -> pd.Series:
    def _normalize_value(value):
        if pd.isna(value):
            return value
        if isinstance(value, str):
            text = value.strip()
            if text.endswith(".0"):
                maybe_int = text[:-2]
                if maybe_int.isdigit():
                    return maybe_int
            return text
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        if isinstance(value, int):
            return str(value)
        text = str(value).strip()
        if text.endswith(".0"):
            maybe_int = text[:-2]
            if maybe_int.isdigit():
                return maybe_int
        return text

    return series.map(_normalize_value)


def render_metric(label: str, value, color_class: str = "blue"):
    st.markdown(f"""
    <div class="metric-card metric-{color_class}">
        <div class="metric-value">{value}</div>
        <div class="metric-label">{label}</div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────
# COLUMN MATCHING
# ─────────────────────────────────────────────────
def find_col(df: pd.DataFrame, candidates: list[str], required: bool = False) -> str | None:
    col_map = {c.strip().lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.strip().lower()
        if key in col_map:
            return col_map[key]
        for k, v in col_map.items():
            if key in k or k in key:
                return v
    if required:
        st.error(f"Could not find column matching any of: {candidates}")
    return None


def clean_columns(df):
    df.columns = [str(c).strip().replace("\n", " ").replace("\r", " ") for c in df.columns]
    return df


def score_sheet_format(df, expected_alias_groups):
    if not expected_alias_groups:
        return 0
    return sum(1 for aliases in expected_alias_groups if find_col(df, aliases) is not None)


@st.cache_data(show_spinner="Loading data…")
def read_upload_from_bytes(file_bytes, file_name, expected_alias_groups, preferred_sheet_keywords):
    name = file_name.lower()
    expected_groups = [list(g) for g in expected_alias_groups]
    preferred_keywords = [k.lower() for k in preferred_sheet_keywords]

    if name.endswith(".csv"):
        df_csv = clean_columns(pd.read_csv(BytesIO(file_bytes)))
        score = score_sheet_format(df_csv, expected_groups)
        return df_csv, "CSV", score, ["CSV"]

    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
    except Exception:
        try:
            df_single = clean_columns(pd.read_excel(BytesIO(file_bytes)))
            score = score_sheet_format(df_single, expected_groups)
            return df_single, "Sheet1", score, []
        except Exception:
            return None, None, 0, []

    best_sheet = None
    best_score = -1
    best_header_row = 0

    for sheet in sheet_names:
        try:
            ws = wb[sheet]
            raw_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                raw_rows.append(list(row))
                if i >= 4:
                    break

            if not raw_rows:
                continue

            best_score_for_sheet = -1
            best_header_for_sheet = 0

            for header_row in [0, 1, 2]:
                if header_row >= len(raw_rows):
                    break
                header_vals = [str(c).strip().replace("\n", " ").replace("\r", " ") if c is not None else f"col_{j}" for j, c in enumerate(raw_rows[header_row])]
                data_rows = raw_rows[header_row + 1:]
                if data_rows:
                    max_len = len(header_vals)
                    padded = [r + [None] * (max_len - len(r)) if len(r) < max_len else r[:max_len] for r in data_rows]
                    mini_df = pd.DataFrame(padded, columns=header_vals)
                else:
                    mini_df = pd.DataFrame(columns=header_vals)

                score = score_sheet_format(mini_df, expected_groups)
                if score > best_score_for_sheet:
                    best_score_for_sheet = score
                    best_header_for_sheet = header_row

            sheet_name_norm = str(sheet).strip().lower().replace("_", " ")
            if any(k in sheet_name_norm for k in preferred_keywords):
                best_score_for_sheet += 1

            if best_score_for_sheet > best_score:
                best_sheet = sheet
                best_score = best_score_for_sheet
                best_header_row = best_header_for_sheet
        except Exception:
            continue

    try:
        wb.close()
    except Exception:
        pass

    if best_sheet is None:
        return None, None, 0, sheet_names

    df_selected = clean_columns(pd.read_excel(
        BytesIO(file_bytes), sheet_name=best_sheet,
        header=best_header_row, engine="openpyxl"
    ))
    return df_selected, best_sheet, best_score, sheet_names


# ─────────────────────────────────────────────────
# COLUMN ALIAS DEFINITIONS
# ─────────────────────────────────────────────────
FC_PROJECT_NAME = ["project name", "project"]
FC_PROJECT_ID = ["project id (icenter)", "project id", "project_id"]
FC_CABINETS_QTY = ["cabinets qty", "cabinet qty", "cabinets quantity", "cab qty"]
FC_NEED_BY_DATE = ["project need by date for shipment (exw)", "need by date", "shipment date"]
FC_FAT_START = ["project need by date for fat start", "fat start date", "fat start"]
FC_ASSEMBLY_START = ["assembly start date", "assembly start"]
FC_ASSEMBLY_END = ["assembly completion date", "assembly end", "assembly completion"]
FC_MATERIAL_AVAIL = ["material availability", "material avail"]
FC_BOM_AVAIL = ["bom availability planned (in oracle)", "bom availability", "bom avail"]
FC_STATUS = ["project status", "status"]
FC_TYPE = ["type (confirmed / forecasted)", "type confirmed forecasted", "confirmed/forecasted", "type"]
FC_SR = ["sr no", "sr", "sr.no", "sr no."]
FC_SHIP_TO = ["ship to"]
FC_BU = ["bu"]
FC_BUILT_BY = ["built by"]
FC_BUILD_PERIOD = ["build period"]
FC_PROMISED_BUILD_DATE = ["promised build date", "promised build datre", "promised build"]
FC_SHIP_PERIOD = ["ship period"]
FC_PROMISED_SHIP = ["promised shipped date", "promised ship date"]
FC_ACTUAL_SHIP = ["actual shipped dates", "actual ship date"]
FC_REMARKS = ["remarks"]

BOM_PROJECT_NUM = ["project num", "project number", "project_num"]
BOM_PROJECT_NAME = ["project name", "project"]
BOM_ORDER_NUM = ["order number", "order num", "order_number"]
BOM_WO_NUM = ["work order number", "work order", "wo number"]
BOM_ITEM = ["item"]
BOM_ITEM_DESC = ["item desc", "item description", "item_desc"]
BOM_COMPONENT_CODE = ["component code", "comp code", "component_code"]
BOM_COMPONENT_DESC = ["component desc", "component description", "comp desc", "component_desc"]
BOM_REQUIRED_QTY = ["required quantity", "required qty", "req qty"]
BOM_ISSUED_QTY = ["quantity issued", "qty issued", "issued qty"]
BOM_OPEN_QTY = ["open quantity2", "open quantity", "open qty"]
BOM_ON_HAND = ["on hand quantity", "on hand qty", "on_hand"]
BOM_INCOMING_PO = ["incoming po quantity", "incoming po qty", "incoming po", "po qty incoming"]
BOM_SCHEDULE_SHIP = ["schedule ship date", "scheduled ship date"]
BOM_MAKE_BUY = ["make or buy", "make/buy"]
BOM_CABINET_TYPE = ["cabinet/buyout/mro", "cabinet buyout mro", "cab/buyout"]
BOM_CUST_PO = ["cust po number", "customer po", "cust po"]
BOM_SALES_STATUS = ["sales status"]
BOM_TOTAL_AVAIL = ["total available", "total avail"]
BOM_PO_RECEIVING = ["po in receiving", "po receiving"]
BOM_SUPPLIER = ["supplier"]
BOM_BUYER = ["buyer name", "buyer", "buyer id", "purchasing buyer", "po buyer"]
BOM_PO_PROMISE = ["po promise date", "promise date", "promised date", "po promised date", "promised dt", "promise dt", "po promise"]
BOM_PO_GENERATED = ["po generated", "po number", "po no", "purchase order", "po #", "po status"]
BOM_JOB_START = ["job start date", "job start"]
BOM_TOTAL_DEMAND = ["total demand"]
BOM_NET_EXT_AVAIL = ["net extended available qty", "net extended avail"]

LOT_NAME_ALIASES = ["lot", "lot name", "lot_name"]
LOT_PROJECT_ALIASES = ["project", "project name", "project_name"]
LOT_ROW_NUM_ALIASES = ["row number", "row no", "row", "rom number", "rom no", "rom nuber", "rom"]


# ─────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>📦 Material Shortage Forecaster</h1>
    <p>Upload Forecasting &amp; Open Orders BOM sheets to predict material shortages across projects</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Data Upload")
    fc_file = st.file_uploader("Forecasting Sheet", type=["xlsx", "xls", "xlsm", "csv"], key="fc")
    bom_file = st.file_uploader("Open Orders BOM Sheet", type=["xlsx", "xls", "xlsm", "csv"], key="bom")

    st.markdown("---")
    st.markdown("### 💾 Saved Lots")
    saved_lots = load_saved_lots()

    if saved_lots:
        lot_names = list(saved_lots.keys())
        selected_saved = st.selectbox("Load a saved lot", ["-- Select --"] + lot_names)
        if selected_saved != "-- Select --":
            st.success(f"Lot **{selected_saved}** loaded")
        col_del1, col_del2 = st.columns(2)
        with col_del1:
            del_lot = st.selectbox("Delete lot", ["-- Select --"] + lot_names, key="del_lot")
        with col_del2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Delete", use_container_width=True):
                if del_lot != "-- Select --":
                    del saved_lots[del_lot]
                    save_lots(saved_lots)
                    st.rerun()
    else:
        st.info("No saved lots yet. Create lots in Project Drill-Down tab.")


# ─────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────
def read_upload(file_obj, expected_alias_groups, preferred_sheet_keywords=()):
    if file_obj is None:
        return None, None, 0, []
    frozen_aliases = tuple(tuple(g) for g in expected_alias_groups)
    frozen_keywords = tuple(preferred_sheet_keywords)
    return read_upload_from_bytes(file_obj.getvalue(), file_obj.name, frozen_aliases, frozen_keywords)


FC_FORMAT_ALIASES = [FC_PROJECT_NAME, FC_CABINETS_QTY, FC_NEED_BY_DATE, FC_STATUS, FC_BUILD_PERIOD]
BOM_FORMAT_ALIASES = [BOM_PROJECT_NAME, BOM_WO_NUM, BOM_ITEM, BOM_COMPONENT_CODE, BOM_REQUIRED_QTY, BOM_OPEN_QTY]

df_fc, fc_sheet_used, fc_format_score, fc_sheets = read_upload(fc_file, FC_FORMAT_ALIASES, ["forecast", "planning", "schedule", "fy"])
df_bom, bom_sheet_used, bom_format_score, bom_sheets = read_upload(bom_file, BOM_FORMAT_ALIASES, ["openordersbom", "open orders bom", "bom", "openorders"])

all_loaded = df_fc is not None and df_bom is not None

if all_loaded:
    with st.sidebar:
        st.markdown("---")
        st.markdown("### 🧭 Sheet Detection")
        st.caption(f"Forecasting: {fc_sheet_used} ({fc_format_score}/{len(FC_FORMAT_ALIASES)} cols)")
        st.caption(f"BOM: {bom_sheet_used} ({bom_format_score}/{len(BOM_FORMAT_ALIASES)} cols)")

if not all_loaded:
    st.info("⬅️ Upload both sheets from the sidebar to begin analysis.")
    with st.expander("ℹ️  Expected Sheet Formats", expanded=True):
        st.markdown("""
**Forecasting Sheet** – Key columns: `Project Name`, `Cabinets Qty`, `Project Need by date for Shipment (EXW)`, `Assembly Start Date`, `Project Status`

**Open Orders BOM Sheet** – Key columns: `Project Name`, `Work Order Number`, `ITEM`, `Item Desc`, `Component Code`, `Required Quantity`, `Open Quantity2`, `On Hand Quantity`, `Incoming PO Quantity`
        """)
    st.stop()


# ─────────────────────────────────────────────────
# RESOLVE COLUMNS
# ─────────────────────────────────────────────────
fc_proj_name_col = find_col(df_fc, FC_PROJECT_NAME)
fc_proj_id_col = find_col(df_fc, FC_PROJECT_ID)
fc_cab_qty_col = find_col(df_fc, FC_CABINETS_QTY)
fc_need_by_col = find_col(df_fc, FC_NEED_BY_DATE)
fc_fat_start_col = find_col(df_fc, FC_FAT_START)
fc_asm_start_col = find_col(df_fc, FC_ASSEMBLY_START)
fc_asm_end_col = find_col(df_fc, FC_ASSEMBLY_END)
fc_mat_avail_col = find_col(df_fc, FC_MATERIAL_AVAIL)
fc_bom_avail_col = find_col(df_fc, FC_BOM_AVAIL)
fc_status_col = find_col(df_fc, FC_STATUS)
fc_type_col = find_col(df_fc, FC_TYPE)
fc_sr_col = find_col(df_fc, FC_SR)
fc_ship_to_col = find_col(df_fc, FC_SHIP_TO)
fc_bu_col = find_col(df_fc, FC_BU)
fc_built_by_col = find_col(df_fc, FC_BUILT_BY)
fc_build_period_col = find_col(df_fc, FC_BUILD_PERIOD)
fc_promised_build_col = find_col(df_fc, FC_PROMISED_BUILD_DATE)
fc_ship_period_col = find_col(df_fc, FC_SHIP_PERIOD)
fc_promised_ship_col = find_col(df_fc, FC_PROMISED_SHIP)
fc_actual_ship_col = find_col(df_fc, FC_ACTUAL_SHIP)
fc_remarks_col = find_col(df_fc, FC_REMARKS)

if not fc_promised_build_col:
    fc_promised_build_col = fc_promised_ship_col

bom_proj_num_col = find_col(df_bom, BOM_PROJECT_NUM)
bom_proj_name_col = find_col(df_bom, BOM_PROJECT_NAME)
bom_order_col = find_col(df_bom, BOM_ORDER_NUM)
bom_wo_col = find_col(df_bom, BOM_WO_NUM)
bom_item_col = find_col(df_bom, BOM_ITEM)
bom_item_desc_col = find_col(df_bom, BOM_ITEM_DESC)
bom_comp_code_col = find_col(df_bom, BOM_COMPONENT_CODE)
bom_comp_desc_col = find_col(df_bom, BOM_COMPONENT_DESC)
bom_req_qty_col = find_col(df_bom, BOM_REQUIRED_QTY)
bom_issued_qty_col = find_col(df_bom, BOM_ISSUED_QTY)
bom_open_qty_col = find_col(df_bom, BOM_OPEN_QTY)
bom_onhand_col = find_col(df_bom, BOM_ON_HAND)
bom_incoming_po_col = find_col(df_bom, BOM_INCOMING_PO)
bom_ship_col = find_col(df_bom, BOM_SCHEDULE_SHIP)
bom_mb_col = find_col(df_bom, BOM_MAKE_BUY)
bom_cab_type_col = find_col(df_bom, BOM_CABINET_TYPE)
bom_cust_po_col = find_col(df_bom, BOM_CUST_PO)
bom_sales_status_col = find_col(df_bom, BOM_SALES_STATUS)
bom_po_recv_col = find_col(df_bom, BOM_PO_RECEIVING)
bom_supplier_col = find_col(df_bom, BOM_SUPPLIER)
bom_buyer_col = find_col(df_bom, BOM_BUYER)
bom_po_promise_col = find_col(df_bom, BOM_PO_PROMISE)
bom_po_generated_col = find_col(df_bom, BOM_PO_GENERATED)
bom_job_start_col = find_col(df_bom, BOM_JOB_START)
bom_total_demand_col = find_col(df_bom, BOM_TOTAL_DEMAND)
bom_net_ext_col = find_col(df_bom, BOM_NET_EXT_AVAIL)


# ─────────────────────────────────────────────────
# PARSE NUMERICS & DATES
# ─────────────────────────────────────────────────
for nc in [fc_cab_qty_col]:
    if nc and nc in df_fc.columns:
        df_fc[nc] = parse_numeric(df_fc[nc])

_fc_date_cols_to_parse = set()
for dc in [fc_need_by_col, fc_fat_start_col, fc_asm_start_col, fc_asm_end_col,
           fc_mat_avail_col, fc_bom_avail_col, fc_promised_build_col, fc_promised_ship_col, fc_actual_ship_col]:
    if dc and dc in df_fc.columns:
        _fc_date_cols_to_parse.add(dc)
for dc in _fc_date_cols_to_parse:
    df_fc[dc] = robust_to_datetime(df_fc[dc])

for nc in [bom_req_qty_col, bom_issued_qty_col, bom_open_qty_col, bom_onhand_col,
           bom_incoming_po_col, bom_po_recv_col, bom_total_demand_col, bom_net_ext_col]:
    if nc and nc in df_bom.columns:
        df_bom[nc] = parse_numeric(df_bom[nc])

for dc in [bom_ship_col, bom_job_start_col]:
    if dc and dc in df_bom.columns:
        df_bom[dc] = robust_to_datetime(df_bom[dc])

for dc in [bom_po_promise_col]:
    if dc and dc in df_bom.columns:
        df_bom[dc] = robust_to_datetime(df_bom[dc])

for id_col in [bom_order_col, bom_wo_col, bom_cust_po_col]:
    if id_col and id_col in df_bom.columns:
        df_bom[id_col] = normalize_identifier_series(df_bom[id_col])


# ─────────────────────────────────────────────────
# COMPUTE TOTAL AVAILABLE FROM BOM (On Hand + Incoming PO + PO in Receiving)
# ─────────────────────────────────────────────────
_bom_oh = df_bom[bom_onhand_col] if bom_onhand_col and bom_onhand_col in df_bom.columns else pd.Series(0, index=df_bom.index)
_bom_po = df_bom[bom_incoming_po_col] if bom_incoming_po_col and bom_incoming_po_col in df_bom.columns else pd.Series(0, index=df_bom.index)
_bom_po_recv = df_bom[bom_po_recv_col] if bom_po_recv_col and bom_po_recv_col in df_bom.columns else pd.Series(0, index=df_bom.index)
df_bom["_total_available"] = _bom_oh + _bom_po + _bom_po_recv

# On-Hand only column for the new table
df_bom["_onhand_only"] = _bom_oh

# Build stock_agg from BOM (component-level max total available)
if bom_comp_code_col:
    stock_agg = (
        df_bom.groupby(bom_comp_code_col, dropna=False)["_total_available"]
        .max().reset_index()
        .rename(columns={bom_comp_code_col: "_stk_item", "_total_available": "_stk_total_onhand"})
    )
    # On-Hand only stock aggregation
    stock_agg_onhand_only = (
        df_bom.groupby(bom_comp_code_col, dropna=False)["_onhand_only"]
        .max().reset_index()
        .rename(columns={bom_comp_code_col: "_stk_item", "_onhand_only": "_stk_total_onhand"})
    )
else:
    stock_agg = pd.DataFrame(columns=["_stk_item", "_stk_total_onhand"])
    stock_agg_onhand_only = pd.DataFrame(columns=["_stk_item", "_stk_total_onhand"])


def compute_priority_shortage_dataframe(df_fc_input, df_bom_input, saved_lots_map, row_lot_assignments_map, stock_agg_override=None):
    """Compute sequentially allocated shortage rows used by multiple tabs.
    
    Args:
        stock_agg_override: If provided, use this stock aggregation instead of the default.
                           Used for On-Hand-Only allocation mode.
    """
    if not bom_comp_code_col or not bom_open_qty_col:
        return pd.DataFrame()

    priority_rows = []
    for _, row in df_fc_input.iterrows():
        rk = row.get("_row_key", "")
        pname = str(row.get(fc_proj_name_col, "")).strip() if pd.notna(row.get(fc_proj_name_col)) else "Unknown"
        asm_date = row.get(fc_asm_start_col) if fc_asm_start_col else None
        bp = str(row.get(fc_build_period_col, "")).strip() if fc_build_period_col and pd.notna(row.get(fc_build_period_col)) else ""
        assigned_lots = [l for l in row_lot_assignments_map.get(rk, []) if l in saved_lots_map]
        if not assigned_lots:
            continue
        for lot_nm in assigned_lots:
            priority_rows.append({
                "row_key": rk,
                "project": pname,
                "lot": lot_nm,
                "build_period": bp,
                "asm_date": asm_date,
                "asm_date_sort": asm_date if pd.notna(asm_date) else pd.Timestamp("2099-12-31"),
            })

    if not priority_rows:
        return pd.DataFrame()

    priority_df = pd.DataFrame(priority_rows).sort_values("asm_date_sort").reset_index(drop=True)
    
    # Use override stock if provided, otherwise use default
    _stock_source = stock_agg_override if stock_agg_override is not None else stock_agg
    remaining_stock = _stock_source.set_index("_stk_item")["_stk_total_onhand"].to_dict()
    all_shortage_rows = []

    for _, prow in priority_df.iterrows():
        lot_nm = prow["lot"]
        lot_info = saved_lots_map.get(lot_nm, {})
        if not lot_info:
            continue

        df_lot_bom = get_lot_bom_scope(df_bom_input, lot_info, bom_proj_name_col, bom_proj_num_col)
        if df_lot_bom.empty:
            continue

        grp = [bom_comp_code_col]
        if bom_comp_desc_col:
            grp.append(bom_comp_desc_col)

        agg_dict = {bom_open_qty_col: "sum"}
        if bom_req_qty_col:
            agg_dict[bom_req_qty_col] = "sum"

        lot_demand = df_lot_bom.groupby(grp, dropna=False).agg(agg_dict).reset_index()

        for _, comp_row in lot_demand.iterrows():
            comp_code = comp_row[bom_comp_code_col]
            comp_desc = comp_row.get(bom_comp_desc_col, "") if bom_comp_desc_col else ""
            open_qty = comp_row[bom_open_qty_col]
            req_qty = comp_row.get(bom_req_qty_col, open_qty) if bom_req_qty_col else open_qty

            available = remaining_stock.get(comp_code, 0)
            allocated = min(available, open_qty)
            shortage = open_qty - allocated
            remaining_stock[comp_code] = available - allocated

            risk = "🔴 Short" if shortage > 0 else "🟢 OK"
            all_shortage_rows.append({
                "Project": prow["project"],
                "Lot": lot_nm,
                "Build Period": prow["build_period"],
                "Assembly Start": prow["asm_date"].strftime("%d-%b-%Y") if pd.notna(prow["asm_date"]) and isinstance(prow["asm_date"], pd.Timestamp) else "N/A",
                "Assembly Start Date": prow["asm_date"],
                "Component Code": comp_code,
                "Component Desc": comp_desc,
                "Required": int(req_qty),
                "Open Qty": int(open_qty),
                "Stock Available (Before)": int(available),
                "Allocated": int(allocated),
                "Shortage": int(shortage),
                "Stock Remaining (After)": int(remaining_stock.get(comp_code, 0)),
                "Risk": risk,
                "_asm_sort": prow["asm_date_sort"],
            })

    if not all_shortage_rows:
        return pd.DataFrame()

    shortage_df = pd.DataFrame(all_shortage_rows)
    shortage_df = shortage_df.sort_values(["_asm_sort", "Project", "Lot", "Shortage"], ascending=[True, True, True, True])
    return shortage_df


# ─────────────────────────────────────────────────
# ROW-LEVEL LOT ASSIGNMENT PERSISTENCE
# ─────────────────────────────────────────────────
ROW_LOT_ASSIGNMENTS_FILE = "row_lot_assignments.json"

def load_row_lot_assignments() -> dict:
    if os.path.exists(ROW_LOT_ASSIGNMENTS_FILE):
        with open(ROW_LOT_ASSIGNMENTS_FILE, "r") as f:
            return json.load(f)
    return {}

def save_row_lot_assignments(assignments: dict):
    with open(ROW_LOT_ASSIGNMENTS_FILE, "w") as f:
        json.dump(assignments, f, indent=2, default=str)


# ─────────────────────────────────────────────────
# ROW KEY HELPERS
# ─────────────────────────────────────────────────
ROW_KEY_SEP = "|||"

def make_fc_row_key(row, proj_col, bp_col, asm_col, idx=None):
    pname = str(row.get(proj_col, "")).strip() if pd.notna(row.get(proj_col)) else ""
    bp = str(row.get(bp_col, "")).strip() if bp_col and pd.notna(row.get(bp_col)) else ""
    asm = ""
    if asm_col and pd.notna(row.get(asm_col)):
        val = row[asm_col]
        asm = val.strftime("%Y-%m-%d") if isinstance(val, pd.Timestamp) else str(val)
    idx_str = str(idx) if idx is not None else ""
    return f"{pname}{ROW_KEY_SEP}{bp}{ROW_KEY_SEP}{asm}{ROW_KEY_SEP}{idx_str}"

def make_fc_row_label(row, proj_col, bp_col, asm_col):
    pname = str(row.get(proj_col, "")).strip() if pd.notna(row.get(proj_col)) else "Unknown"
    bp = str(row.get(bp_col, "")).strip() if bp_col and pd.notna(row.get(bp_col)) else "N/A"
    asm = "N/A"
    if asm_col and pd.notna(row.get(asm_col)):
        val = row[asm_col]
        asm = val.strftime("%d-%b-%Y") if isinstance(val, pd.Timestamp) else str(val)
    return f"{pname}  |  Build: {bp}  |  Asm Start: {asm}"


# ─────────────────────────────────────────────────
# ADD ROW KEYS TO FORECASTING DF
# ─────────────────────────────────────────────────
df_fc["_row_key"] = [
    make_fc_row_key(row, fc_proj_name_col, fc_build_period_col, fc_asm_start_col, idx=i)
    for i, row in df_fc.iterrows()
]
df_fc["_row_label"] = [
    make_fc_row_label(row, fc_proj_name_col, fc_build_period_col, fc_asm_start_col)
    for _, row in df_fc.iterrows()
]


# ─────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────
tab_project, tab_forecasting, tab_shortage, tab_analytics, tab_timeline = st.tabs([
    "🏗️ Project Drill-Down", "📈 Forecasting",
    "🚨 Shortage Forecast", "📊 Lot & Project Analytics", "📅 Timeline"
])


# ═══════════════════════════════════════════════════
# TAB 1 – PROJECT DRILL-DOWN
# ═══════════════════════════════════════════════════
with tab_project:
    st.markdown('<div class="section-header">Project → Work Order → Item → Component Drill-Down</div>', unsafe_allow_html=True)

    project_list = sorted(df_bom[bom_proj_name_col].dropna().unique().tolist()) if bom_proj_name_col else []
    if not project_list:
        project_list = sorted(df_bom[bom_proj_num_col].dropna().unique().tolist()) if bom_proj_num_col else []

    selected_project = st.selectbox("Select Project", ["-- All --"] + project_list, key="proj_sel")
    df_filtered = df_bom.copy()
    if selected_project != "-- All --":
        proj_col_used = bom_proj_name_col or bom_proj_num_col
        if proj_col_used:
            df_filtered = df_filtered[df_filtered[proj_col_used] == selected_project]

    selected_wos = []
    selected_item_codes = []
    selected_item_labels = []

    wo_list = sorted(df_filtered[bom_wo_col].dropna().unique().tolist()) if bom_wo_col else []
    wo_options = [str(w) for w in wo_list]
    selected_wos = st.multiselect("Select Work Orders", wo_options, default=[], key="wo_sel_multi", help="Leave empty = all")
    if selected_wos and bom_wo_col:
        df_filtered = df_filtered[df_filtered[bom_wo_col].astype(str).isin(selected_wos)]

    if bom_item_col and bom_item_desc_col:
        item_combos = df_filtered[[bom_item_col, bom_item_desc_col]].drop_duplicates().dropna(subset=[bom_item_col])
        item_combos["_label"] = item_combos[bom_item_col].astype(str) + " — " + item_combos[bom_item_desc_col].fillna("").astype(str)
        item_labels = sorted(item_combos["_label"].tolist())
        selected_item_labels = st.multiselect("Select ITEM Codes", item_labels, default=[], key="item_sel_multi", help="Leave empty = all")
        if selected_item_labels:
            selected_item_codes = [lbl.split(" — ")[0].strip() for lbl in selected_item_labels]
            df_filtered = df_filtered[df_filtered[bom_item_col].astype(str).isin(selected_item_codes)]
    elif bom_item_col:
        item_list = sorted(df_filtered[bom_item_col].dropna().unique().tolist())
        selected_item_codes = st.multiselect("Select ITEM Codes", [str(i) for i in item_list], default=[], key="item_sel_multi_plain")
        if selected_item_codes:
            df_filtered = df_filtered[df_filtered[bom_item_col].astype(str).isin(selected_item_codes)]

    scope_wo_count = len(selected_wos) if selected_wos else len(wo_options)
    scope_item_count = len(selected_item_codes) if selected_item_codes else (df_filtered[bom_item_col].dropna().astype(str).nunique() if bom_item_col else 0)
    st.caption(f"Scope: {scope_wo_count} Work Order(s), {scope_item_count} ITEM(s).")

    st.markdown('<div class="section-header">Component Detail</div>', unsafe_allow_html=True)

    display_cols = [c for c in [
        bom_comp_code_col, bom_comp_desc_col, bom_req_qty_col, bom_issued_qty_col,
        bom_open_qty_col, bom_onhand_col, bom_incoming_po_col,
        bom_po_recv_col, bom_supplier_col
    ] if c is not None]

    display_cols_with_ta = list(dict.fromkeys(display_cols + ["_total_available"]))

    comp_detail = pd.DataFrame()
    if display_cols:
        available_cols = [c for c in display_cols_with_ta if c in df_filtered.columns]
        comp_base = df_filtered[available_cols].copy()
        if bom_comp_code_col:
            group_keys = [bom_comp_code_col]
            if bom_comp_desc_col:
                group_keys.append(bom_comp_desc_col)
            agg_map = {}
            for sc in [bom_req_qty_col, bom_issued_qty_col, bom_open_qty_col]:
                if sc and sc in comp_base.columns and sc not in group_keys:
                    agg_map[sc] = "sum"
            for mc in [bom_onhand_col, bom_incoming_po_col, bom_po_recv_col, "_total_available"]:
                if mc and mc in comp_base.columns and mc not in group_keys:
                    agg_map[mc] = "max"
            for fc_c in [bom_supplier_col]:
                if fc_c and fc_c in comp_base.columns and fc_c not in group_keys and fc_c not in agg_map:
                    agg_map[fc_c] = "first"
            comp_detail = comp_base.groupby(group_keys, dropna=False, as_index=False).agg(agg_map)
            if bom_wo_col and bom_wo_col in df_filtered.columns:
                wo_counts = df_filtered.groupby(group_keys, dropna=False)[bom_wo_col].nunique().reset_index(name="Work Orders")
                comp_detail = comp_detail.merge(wo_counts, on=group_keys, how="left")
        else:
            comp_detail = comp_base.copy()

        if "_total_available" in comp_detail.columns and bom_open_qty_col and bom_open_qty_col in comp_detail.columns:
            comp_detail["Total Available vs Open Qty"] = comp_detail["_total_available"] - comp_detail[bom_open_qty_col]

        comp_detail.rename(columns={"_total_available": "Total Available (OH+PO+Recv)"}, inplace=True)

        if bom_open_qty_col and bom_open_qty_col in comp_detail.columns:
            comp_detail = comp_detail.sort_values(by=bom_open_qty_col, ascending=False)

        st.dataframe(comp_detail, use_container_width=True, height=450)

        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1: render_metric("Components", len(comp_detail), "cyan")
        with mc2:
            if bom_open_qty_col:
                render_metric("Total Open Qty", int(comp_detail[bom_open_qty_col].sum()), "amber")
        with mc3:
            if "Total Available vs Open Qty" in comp_detail.columns:
                render_metric("Short Components", int((comp_detail["Total Available vs Open Qty"] < 0).sum()), "red")
        with mc4:
            if bom_req_qty_col:
                render_metric("Total Required", int(comp_detail[bom_req_qty_col].sum()), "purple")

    st.markdown("---")
    st.markdown("### 📥 Upload Lot Details (Excel/CSV)")
    lot_upload_file = st.file_uploader("Upload Lot File", type=["xlsx", "xls", "xlsm", "csv"], key="lot_upload_file")
    if lot_upload_file is not None:
        lot_df, lot_sheet, lot_score, _ = read_upload(
            lot_upload_file,
            [LOT_NAME_ALIASES, LOT_PROJECT_ALIASES, LOT_ROW_NUM_ALIASES],
            ["lot", "mapping", "details"],
        )
        if lot_df is None:
            st.error("Unable to read lot upload file.")
        else:
            lot_name_col = find_col(lot_df, LOT_NAME_ALIASES)
            lot_project_col = find_col(lot_df, LOT_PROJECT_ALIASES)
            lot_row_col = find_col(lot_df, LOT_ROW_NUM_ALIASES)
            st.caption(f"Detected sheet: {lot_sheet} ({lot_score}/3 key columns matched)")

            if not lot_name_col or not lot_project_col:
                st.error("Lot file must contain at least Lot Name and Project Name columns.")
            else:
                preview_cols = [c for c in [lot_name_col, lot_project_col, lot_row_col] if c]
                st.dataframe(lot_df[preview_cols].head(20), use_container_width=True, height=250)

                if st.button("Import Lots from File", key="import_lots_from_file"):
                    saved_lots_current = load_saved_lots()
                    imported_count = 0
                    skipped_count = 0
                    conflict_rows = []

                    for idx, lot_row in lot_df.iterrows():
                        lot_name_val = str(lot_row.get(lot_name_col, "")).strip() if pd.notna(lot_row.get(lot_name_col)) else ""
                        project_val = str(lot_row.get(lot_project_col, "")).strip() if pd.notna(lot_row.get(lot_project_col)) else ""
                        row_num_val = ""
                        if lot_row_col and pd.notna(lot_row.get(lot_row_col)):
                            row_num_val = str(lot_row.get(lot_row_col)).strip()

                        if not lot_name_val or not project_val:
                            skipped_count += 1
                            continue

                        has_conflict = False
                        for existing_lot_name, existing_info in saved_lots_current.items():
                            if existing_lot_name == lot_name_val:
                                continue
                            same_project = get_lot_project_name(existing_info) == project_val
                            same_row = get_lot_row_number(existing_info) == row_num_val
                            if same_project and row_num_val and same_row:
                                has_conflict = True
                                conflict_rows.append(f"Row {idx + 1}: {lot_name_val} conflicts with existing lot {existing_lot_name}")
                                break

                        if has_conflict:
                            skipped_count += 1
                            continue

                        saved_lots_current[lot_name_val] = {
                            "project_name": project_val,
                            "row_number": row_num_val,
                        }
                        imported_count += 1

                    save_lots(saved_lots_current)
                    if imported_count:
                        st.success(f"Imported/updated {imported_count} lot(s).")
                    if skipped_count:
                        st.warning(f"Skipped {skipped_count} row(s) due to missing data or conflicts.")
                    if conflict_rows:
                        st.caption("Conflicts: " + " | ".join(conflict_rows[:5]))
                    st.rerun()

    st.markdown("---")
    st.markdown("### 💾 Save Current Selection as Lot")
    lot_name = st.text_input("Lot Name", key="lot_name_input")
    lot_row_number = st.text_input("ROW / ROM Number", key="lot_row_number_input")
    if st.button("Save Lot", type="primary", key="save_lot_btn"):
        if lot_name.strip():
            lot_name_clean = lot_name.strip()
            lot_row_clean = lot_row_number.strip()
            project_clean = str(selected_project).strip()

            saved_lots_current = load_saved_lots()
            validation_errors = []

            if project_clean == "-- All --" or not project_clean:
                validation_errors.append("❌ Select a specific Project before saving the lot.")
            if not lot_row_clean:
                validation_errors.append("❌ Enter ROW / ROM Number.")

            for existing_lot_name, existing_lot_info in saved_lots_current.items():
                if existing_lot_name == lot_name_clean:
                    continue
                same_project = get_lot_project_name(existing_lot_info) == project_clean
                same_row = get_lot_row_number(existing_lot_info) == lot_row_clean
                if same_project and same_row and lot_row_clean:
                    validation_errors.append(
                        f"❌ ROW / ROM Number **{lot_row_clean}** already exists for project **{project_clean}** in lot **{existing_lot_name}**"
                    )
                    break
            
            if validation_errors:
                st.error("⚠️ **Cannot save lot due to conflicts:**\n\n" + "\n\n".join(validation_errors))
            else:
                lot_info = {
                    "project_name": project_clean,
                    "row_number": lot_row_clean,
                }
                saved_lots_current[lot_name_clean] = lot_info
                save_lots(saved_lots_current)
                st.success(f"✅ Lot **{lot_name_clean}** saved successfully!")
                st.rerun()
        else:
            st.error("Enter a lot name.")


# ═══════════════════════════════════════════════════
# TAB 2 – FORECASTING (with Row-Level Lot Assignment)
# ═══════════════════════════════════════════════════
with tab_forecasting:
    st.markdown('<div class="section-header">📈 Forecasting Schedule with Row-Level Lot Assignment</div>', unsafe_allow_html=True)

    if not fc_proj_name_col:
        st.error("❌ Project Name column not found in Forecasting sheet.")
        st.stop()
    else:
        fc_display_cols = []
        seen_cols = set()
        for c in [fc_sr_col, fc_proj_name_col, fc_proj_id_col, fc_ship_to_col, fc_bu_col, fc_type_col,
                  fc_cab_qty_col, fc_built_by_col, fc_need_by_col, fc_fat_start_col, fc_build_period_col,
                  fc_promised_build_col, fc_bom_avail_col, fc_mat_avail_col, fc_asm_start_col, fc_asm_end_col,
                  fc_ship_period_col, fc_promised_ship_col, fc_actual_ship_col, fc_status_col, fc_remarks_col]:
            if c is not None and c in df_fc.columns and c not in seen_cols:
                fc_display_cols.append(c)
                seen_cols.add(c)
        if not fc_display_cols:
            fc_display_cols = [fc_proj_name_col]

        fc_view = df_fc[fc_display_cols + ["_row_key", "_row_label"]].copy()
        for col in fc_view.columns:
            if col not in ["_row_key", "_row_label"] and pd.api.types.is_datetime64_any_dtype(fc_view[col]):
                fc_view[col] = fc_view[col].dt.strftime("%d-%b-%Y").fillna("")
        fc_view = fc_view.dropna(how='all', subset=[c for c in fc_display_cols if c in fc_view.columns])

        # ── ROW-LEVEL Lot Assignment ──
        st.markdown('<div class="section-header">🔗 Assign Lots to Forecasting Schedule Rows</div>', unsafe_allow_html=True)
        st.caption("Assign lots to specific rows identified by Project Name + Build Period + Assembly Start Date")

        row_lot_assignments = load_row_lot_assignments()
        saved_lots = load_saved_lots()

        if not saved_lots:
            st.info("ℹ️  No saved lots yet. Create lots in the **Project Drill-Down** tab first.")
        else:
            lot_name_options = sorted(saved_lots.keys())
            row_options_map = {}
            for _, row in fc_view.iterrows():
                row_options_map[row["_row_label"]] = row["_row_key"]
            row_labels_sorted = sorted(row_options_map.keys())

            ac1, ac2, ac3 = st.columns([3, 2, 1])
            with ac1:
                assign_row_label = st.selectbox("Select Forecasting Row (Project | Build Period | Asm Start)", ["-- Select --"] + row_labels_sorted, key="fc_assign_row")
            with ac2:
                assign_lots = st.multiselect("Assign Lots", lot_name_options, key="fc_assign_lots")
            with ac3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("✅ Assign", use_container_width=True, key="fc_assign_btn"):
                    if assign_row_label != "-- Select --" and assign_lots:
                        row_key = row_options_map.get(assign_row_label, "")
                        if row_key:
                            existing = row_lot_assignments.get(row_key, [])
                            merged = list(dict.fromkeys(existing + assign_lots))
                            row_lot_assignments[row_key] = merged
                            save_row_lot_assignments(row_lot_assignments)
                            st.success(f"Assigned {len(assign_lots)} lot(s) to row")
                            st.rerun()
                    else:
                        st.error("Select both a row and at least one lot.")

            if row_lot_assignments:
                with st.expander("📋 Current Row-Level Lot Assignments", expanded=False):
                    key_to_label = {v: k for k, v in row_options_map.items()}
                    for rk, lots_list in sorted(row_lot_assignments.items(), key=lambda x: x[0]):
                        if lots_list:
                            display_label = key_to_label.get(rk, rk)
                            chips = " ".join([f'<span class="lot-chip">{l}</span>' for l in lots_list])
                            st.markdown(f"**{display_label}:** {chips}", unsafe_allow_html=True)
                    st.markdown("---")
                    rc1, rc2, rc3 = st.columns([3, 2, 1])
                    with rc1:
                        assigned_labels = [key_to_label.get(k, k) for k in row_lot_assignments.keys() if row_lot_assignments[k]]
                        rm_row_label = st.selectbox("Remove from row", ["-- Select --"] + sorted(assigned_labels), key="fc_rm_row")
                    with rc2:
                        rm_row_key = row_options_map.get(rm_row_label, "") if rm_row_label != "-- Select --" else ""
                        rm_options = row_lot_assignments.get(rm_row_key, [])
                        rm_lots = st.multiselect("Lots to remove", rm_options, key="fc_rm_lots")
                    with rc3:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("🗑️ Remove", use_container_width=True, key="fc_rm_btn"):
                            if rm_row_key and rm_lots:
                                row_lot_assignments[rm_row_key] = [l for l in row_lot_assignments.get(rm_row_key, []) if l not in rm_lots]
                                if not row_lot_assignments[rm_row_key]:
                                    del row_lot_assignments[rm_row_key]
                                save_row_lot_assignments(row_lot_assignments)
                                st.rerun()

        # ── Enrich table with lot info per row ──
        lot_col_assigned, lot_col_comp_count, lot_col_short, lot_col_open_qty = [], [], [], []
        lot_metrics_cache = {}
        for _, row in fc_view.iterrows():
            rk = row["_row_key"]
            assigned = [l for l in row_lot_assignments.get(rk, []) if l in saved_lots]
            if assigned:
                lot_col_assigned.append(", ".join(assigned))
                comp_total = 0
                short_total = 0
                open_total = 0
                for lot_name in assigned:
                    if lot_name not in lot_metrics_cache:
                        lot_metrics_cache[lot_name] = compute_lot_metrics(df_bom, saved_lots.get(lot_name, {}))
                    metrics = lot_metrics_cache[lot_name]
                    comp_total += metrics["component_count"]
                    short_total += metrics["short_components"]
                    open_total += metrics["total_open_qty"]
                lot_col_comp_count.append(comp_total)
                lot_col_short.append(short_total)
                lot_col_open_qty.append(open_total)
            else:
                lot_col_assigned.append("")
                lot_col_comp_count.append(0)
                lot_col_short.append(0)
                lot_col_open_qty.append(0)

        fc_view["Assigned Lots"] = lot_col_assigned
        fc_view["Lot Components"] = lot_col_comp_count
        fc_view["Lot Shortages"] = lot_col_short
        fc_view["Lot Open Qty"] = lot_col_open_qty

        if fc_asm_start_col and fc_asm_start_col in df_fc.columns:
            sort_order = df_fc[fc_asm_start_col].sort_values(na_position="last").index
            valid_idx = [i for i in sort_order if i in fc_view.index]
            fc_view = fc_view.loc[valid_idx].reset_index(drop=True)

        # Filters
        flt1, flt2, flt3 = st.columns(3)
        with flt1:
            fc_filter_status = []
            if fc_status_col and fc_status_col in fc_view.columns:
                statuses = sorted(fc_view[fc_status_col].dropna().unique().tolist())
                fc_filter_status = st.multiselect("Filter by Status", statuses, key="fc_flt_status")
        with flt2:
            fc_filter_lot = st.selectbox("Lot Filter", ["All", "With Lots Only", "Without Lots Only"], key="fc_flt_lot")
        with flt3:
            fc_search = st.text_input("Search Project", key="fc_search_proj")

        fc_display = fc_view.copy()
        if fc_filter_status and fc_status_col in fc_display.columns:
            fc_display = fc_display[fc_display[fc_status_col].isin(fc_filter_status)]
        if fc_filter_lot == "With Lots Only":
            fc_display = fc_display[fc_display["Assigned Lots"] != ""]
        elif fc_filter_lot == "Without Lots Only":
            fc_display = fc_display[fc_display["Assigned Lots"] == ""]
        if fc_search:
            fc_display = fc_display[fc_display[fc_proj_name_col].astype(str).str.contains(fc_search, case=False, na=False)]

        st.dataframe(fc_display.drop(columns=["_row_key", "_row_label"], errors="ignore"), use_container_width=True, height=550)
        st.caption(f"Showing {len(fc_display):,} of {len(fc_view):,} rows.")

        # Lot detail drill-down per row
        rows_with_lots = fc_display[fc_display["Assigned Lots"] != ""]
        if not rows_with_lots.empty:
            st.markdown('<div class="section-header">Lot Component Detail per Row</div>', unsafe_allow_html=True)
            detail_row_labels = rows_with_lots["_row_label"].tolist()
            detail_row_sel = st.selectbox("Select row to view lot details", ["-- Select --"] + detail_row_labels, key="fc_lot_detail_row")
            if detail_row_sel != "-- Select --":
                sel_row = rows_with_lots[rows_with_lots["_row_label"] == detail_row_sel].iloc[0]
                sel_rk = sel_row["_row_key"]
                assigned_lot_names = [l for l in row_lot_assignments.get(sel_rk, []) if l in saved_lots]
                for lot_nm in assigned_lot_names:
                    lot_info = saved_lots.get(lot_nm, {})
                    if not lot_info:
                        continue
                    lot_metrics = compute_lot_metrics(df_bom, lot_info)
                    with st.expander(f"📦 Lot: **{lot_nm}**", expanded=True):
                        d1, d2, d3, d4 = st.columns(4)
                        with d1: st.metric("Project", get_lot_project_name(lot_info) or "N/A")
                        with d2: st.metric("ROW / ROM", get_lot_row_number(lot_info) or "N/A")
                        with d3: st.metric("Components", lot_metrics["component_count"])
                        with d4: st.metric("Short", lot_metrics["short_components"])

                        df_lot_bom = get_lot_bom_scope(df_bom, lot_info, bom_proj_name_col, bom_proj_num_col)

                        if not df_lot_bom.empty and bom_comp_code_col and bom_open_qty_col:
                            grp = [bom_comp_code_col]
                            if bom_comp_desc_col:
                                grp.append(bom_comp_desc_col)
                            lot_comp = df_lot_bom.groupby(grp, dropna=False).agg(
                                Open_Qty=(bom_open_qty_col, "sum"),
                                Total_Available=("_total_available", "max"),
                            ).reset_index()
                            lot_comp["Gap"] = lot_comp["Total_Available"] - lot_comp["Open_Qty"]
                            lot_comp = lot_comp.sort_values("Gap", ascending=True)
                            st.dataframe(lot_comp, use_container_width=True, height=300)
                        else:
                            st.caption("No matching BOM data for this lot.")

        st.markdown('<div class="section-header">📆 Month-wise Schedule + Lot Assignment Analytics</div>', unsafe_allow_html=True)

        if fc_asm_start_col and fc_asm_start_col in df_fc.columns:
            month_rows = []
            month_proj_lot_rows = []
            for _, frow in df_fc.iterrows():
                asm_dt = frow.get(fc_asm_start_col)
                if pd.isna(asm_dt):
                    continue

                month_key = pd.Timestamp(asm_dt).to_period("M").to_timestamp()
                project_name = str(frow.get(fc_proj_name_col, "Unknown")).strip() if pd.notna(frow.get(fc_proj_name_col)) else "Unknown"
                row_key = frow.get("_row_key", "")
                assigned_lots = [l for l in row_lot_assignments.get(row_key, []) if l in saved_lots]

                month_rows.append({
                    "Month": month_key,
                    "Project": project_name,
                    "Row Key": row_key,
                    "Has Lots": len(assigned_lots) > 0,
                })

                for lot_nm in assigned_lots:
                    month_proj_lot_rows.append({
                        "Month": month_key,
                        "Project": project_name,
                        "Lot": lot_nm,
                        "Project-Lot": f"{project_name} → {lot_nm}",
                    })

            if month_rows:
                month_df = pd.DataFrame(month_rows)
                month_summary = month_df.groupby("Month", as_index=False).agg(
                    Projects_Scheduled=("Project", "nunique"),
                    Rows_Scheduled=("Row Key", "nunique"),
                    Rows_With_Lots=("Has Lots", "sum"),
                )

                if month_proj_lot_rows:
                    month_proj_lot_df = pd.DataFrame(month_proj_lot_rows)
                    month_lot_summary = month_proj_lot_df.groupby("Month", as_index=False).agg(
                        Unique_Lots_Assigned=("Lot", "nunique"),
                        Project_Lot_Pairs=("Project-Lot", "nunique"),
                    )
                    month_summary = month_summary.merge(month_lot_summary, on="Month", how="left")
                else:
                    month_summary["Unique_Lots_Assigned"] = 0
                    month_summary["Project_Lot_Pairs"] = 0

                month_summary = month_summary.sort_values("Month")
                month_summary_display = month_summary.copy()
                month_summary_display["Month"] = month_summary_display["Month"].dt.strftime("%b-%Y")

                st.dataframe(month_summary_display, use_container_width=True, height=250)

                fig_month = go.Figure()
                fig_month.add_trace(go.Bar(
                    x=month_summary_display["Month"],
                    y=month_summary_display["Projects_Scheduled"],
                    name="Projects",
                    marker_color="#3b82f6",
                    text=month_summary_display["Projects_Scheduled"],
                    textposition="outside",
                    textfont=dict(color="#e2e8f0", size=12, family="DM Sans"),
                ))
                fig_month.add_trace(go.Bar(
                    x=month_summary_display["Month"],
                    y=month_summary_display["Unique_Lots_Assigned"],
                    name="Unique Lots",
                    marker_color="#8b5cf6",
                    text=month_summary_display["Unique_Lots_Assigned"],
                    textposition="outside",
                    textfont=dict(color="#e2e8f0", size=12, family="DM Sans"),
                ))
                fig_month.update_layout(
                    barmode="group",
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8"),
                    margin=dict(l=10, r=10, t=40, b=30),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.2),
                    xaxis_title="",
                    yaxis_title="Count"
                )
                st.plotly_chart(fig_month, use_container_width=True)

                st.markdown("#### Month-wise Project + Lot Analytics")
                if month_proj_lot_rows:
                    seq_shortage_df = compute_priority_shortage_dataframe(df_fc, df_bom, saved_lots, row_lot_assignments)
                    month_proj_lot_view = pd.DataFrame(month_proj_lot_rows)
                    month_proj_lot_view = month_proj_lot_view.groupby("Month", as_index=False).agg(
                        Project_Lot_Pairs=("Project-Lot", "nunique")
                    )

                    if not seq_shortage_df.empty:
                        seq_shortage_df = seq_shortage_df.copy()
                        seq_shortage_df["Month"] = pd.to_datetime(seq_shortage_df["Assembly Start Date"], errors="coerce").dt.to_period("M").dt.to_timestamp()
                        seq_shortage_df = seq_shortage_df.dropna(subset=["Month"])

                        month_short = seq_shortage_df.groupby("Month", as_index=False).agg(
                            Total_Unique_Components=("Component Code", "nunique")
                        )
                        month_short_only = (
                            seq_shortage_df[seq_shortage_df["Shortage"] > 0]
                            .groupby("Month", as_index=False)
                            .agg(Shortage_Unique_Components=("Component Code", "nunique"))
                        )
                        month_short = month_short.merge(month_short_only, on="Month", how="left")
                        month_short["Shortage_Unique_Components"] = month_short["Shortage_Unique_Components"].fillna(0).astype(int)
                        month_short["Shortage %"] = (
                            month_short["Shortage_Unique_Components"]
                            / month_short["Total_Unique_Components"].replace(0, pd.NA)
                            * 100
                        ).fillna(0)

                        month_proj_lot_view = month_proj_lot_view.merge(month_short, on="Month", how="left")
                    else:
                        month_proj_lot_view["Total_Unique_Components"] = 0
                        month_proj_lot_view["Shortage_Unique_Components"] = 0
                        month_proj_lot_view["Shortage %"] = 0.0

                    month_proj_lot_view = month_proj_lot_view.sort_values("Month")
                    month_proj_lot_display = month_proj_lot_view.copy()
                    month_proj_lot_display["Month"] = month_proj_lot_display["Month"].dt.strftime("%b-%Y")
                    month_proj_lot_display["Shortage %"] = month_proj_lot_display["Shortage %"].map(lambda x: f"{x:.1f}%")
                    st.dataframe(month_proj_lot_display, use_container_width=True, height=240)
                else:
                    st.caption("No lot assignments found for month-wise project + lot analytics.")
            else:
                st.caption("Assembly Start Date is required for month-wise analytics.")
        else:
            st.caption("Assembly Start Date column not found; month-wise analytics unavailable.")


# ═══════════════════════════════════════════════════
# TAB 3 – SHORTAGE FORECAST (Priority-based Stock Allocation)
# ═══════════════════════════════════════════════════
with tab_shortage:
    st.markdown('<div class="section-header">Priority-Based Material Shortage Forecast (Project + Lot)</div>', unsafe_allow_html=True)
    st.caption("Total Available (On Hand + Incoming PO + PO in Receiving) is allocated to projects in order of Assembly Start Date. Earlier projects get stock first.")

    row_lot_assignments = load_row_lot_assignments()
    saved_lots = load_saved_lots()
    shortage_df = compute_priority_shortage_dataframe(df_fc, df_bom, saved_lots, row_lot_assignments)

    if bom_comp_code_col and bom_open_qty_col:
        if shortage_df.empty:
            st.info("ℹ️  No lots assigned to forecasting rows yet. Assign lots in the **Forecasting** tab first.")
        else:
            shortage_display = shortage_df.drop(columns=["_asm_sort"], errors="ignore")

            f1, f2, f3 = st.columns(3)
            with f1:
                risk_filter = st.multiselect("Shortage Status", ["🔴 Short", "🟢 OK"], default=["🔴 Short"], key="risk_flt")
            with f2:
                proj_flt = st.multiselect("Projects", sorted(shortage_display["Project"].unique().tolist()), key="proj_flt")
            with f3:
                lot_flt = st.multiselect("Lots", sorted(shortage_display["Lot"].unique().tolist()), key="lot_flt")

            base_filtered = shortage_display.copy()
            if proj_flt:
                base_filtered = base_filtered[base_filtered["Project"].isin(proj_flt)]
            if lot_flt:
                base_filtered = base_filtered[base_filtered["Lot"].isin(lot_flt)]

            filtered = base_filtered[base_filtered["Risk"].isin(risk_filter)].copy()

            short_comp = base_filtered[base_filtered["Risk"] == "🔴 Short"]["Component Code"].nunique()
            ok_comp = base_filtered[base_filtered["Risk"] == "🟢 OK"]["Component Code"].nunique()
            shortage_unique = base_filtered[base_filtered["Shortage"] > 0]["Component Code"].nunique()
            total_unique = base_filtered["Component Code"].nunique()
            shortage_pct = (shortage_unique / total_unique * 100) if total_unique else 0

            mc1, mc2, mc3, mc4 = st.columns(4)
            with mc1: render_metric("Short Components", short_comp, "red")
            with mc2: render_metric("OK Components", ok_comp, "emerald")
            with mc3: render_metric("Total Unique Components", total_unique, "cyan")
            with mc4: render_metric("Shortage %", f"{shortage_pct:.1f}%", "purple")

            st.dataframe(filtered, use_container_width=True, height=500)

            st.markdown('<div class="section-header">Project + Lot Shortage Status Heatmap</div>', unsafe_allow_html=True)
            risk_summary = shortage_display.copy()
            risk_summary["Proj_Lot"] = risk_summary["Project"] + " → " + risk_summary["Lot"]
            risk_heatmap = risk_summary.groupby("Proj_Lot").agg(
                Short=("Risk", lambda x: (x == "🔴 Short").sum()),
                OK=("Risk", lambda x: (x == "🟢 OK").sum()),
            ).reset_index().sort_values("Short", ascending=False)
            risk_heatmap["Total"] = risk_heatmap["Short"] + risk_heatmap["OK"]
            risk_heatmap["Short %"] = (risk_heatmap["Short"] / risk_heatmap["Total"].replace(0, 1) * 100).round(1)
            risk_heatmap["OK %"] = (risk_heatmap["OK"] / risk_heatmap["Total"].replace(0, 1) * 100).round(1)

            fig_heat = go.Figure()
            fig_heat.add_trace(go.Bar(
                y=risk_heatmap["Proj_Lot"], x=risk_heatmap["Short"], name="Short",
                marker_color="#ef4444", orientation="h",
                text=risk_heatmap["Short %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside", textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_heat.add_trace(go.Bar(
                y=risk_heatmap["Proj_Lot"], x=risk_heatmap["OK"], name="OK",
                marker_color="#10b981", orientation="h",
                text=risk_heatmap["OK %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside", textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_heat.update_layout(barmode="stack", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8"), height=max(300, len(risk_heatmap) * 35),
                margin=dict(l=10, r=10, t=10, b=30), legend=dict(orientation="h", yanchor="bottom", y=-0.15))
            st.plotly_chart(fig_heat, use_container_width=True)

            # ───────────────────────────────────────────────
            # NEW: On-Hand Only Shortage Allocation Table
            # ───────────────────────────────────────────────
            st.markdown('<div class="section-header">🏭 On-Hand Only Shortage Allocation</div>', unsafe_allow_html=True)
            st.caption("Allocation based **only on On Hand Quantity** (excludes Incoming PO and PO in Receiving). Compared against the full allocation above to highlight the gap that PO stock covers.")

            shortage_oh_df = compute_priority_shortage_dataframe(
                df_fc, df_bom, saved_lots, row_lot_assignments,
                stock_agg_override=stock_agg_onhand_only
            )

            if shortage_oh_df.empty:
                st.info("No data for On-Hand Only allocation.")
            else:
                shortage_oh_display = shortage_oh_df.drop(columns=["_asm_sort"], errors="ignore")

                # Apply same project/lot filters
                oh_filtered = shortage_oh_display.copy()
                if proj_flt:
                    oh_filtered = oh_filtered[oh_filtered["Project"].isin(proj_flt)]
                if lot_flt:
                    oh_filtered = oh_filtered[oh_filtered["Lot"].isin(lot_flt)]

                oh_base = oh_filtered.copy()
                oh_short_comp = oh_base[oh_base["Risk"] == "🔴 Short"]["Component Code"].nunique()
                oh_ok_comp = oh_base[oh_base["Risk"] == "🟢 OK"]["Component Code"].nunique()
                oh_total_unique = oh_base["Component Code"].nunique()
                oh_shortage_pct = (oh_short_comp / oh_total_unique * 100) if oh_total_unique else 0

                om1, om2, om3, om4 = st.columns(4)
                with om1: render_metric("Short (OH Only)", oh_short_comp, "red")
                with om2: render_metric("OK (OH Only)", oh_ok_comp, "emerald")
                with om3: render_metric("Total Components", oh_total_unique, "cyan")
                with om4: render_metric("Shortage % (OH Only)", f"{oh_shortage_pct:.1f}%", "purple")

                # Filter by risk if selected
                oh_risk_filtered = oh_base[oh_base["Risk"].isin(risk_filter)].copy()
                st.dataframe(oh_risk_filtered, use_container_width=True, height=500)

                # On-Hand Only heatmap
                st.markdown('<div class="section-header">On-Hand Only – Project + Lot Heatmap</div>', unsafe_allow_html=True)
                oh_risk_summary = shortage_oh_display.copy()
                oh_risk_summary["Proj_Lot"] = oh_risk_summary["Project"] + " → " + oh_risk_summary["Lot"]
                oh_risk_heatmap = oh_risk_summary.groupby("Proj_Lot").agg(
                    Short=("Risk", lambda x: (x == "🔴 Short").sum()),
                    OK=("Risk", lambda x: (x == "🟢 OK").sum()),
                ).reset_index().sort_values("Short", ascending=False)
                oh_risk_heatmap["Total"] = oh_risk_heatmap["Short"] + oh_risk_heatmap["OK"]
                oh_risk_heatmap["Short %"] = (oh_risk_heatmap["Short"] / oh_risk_heatmap["Total"].replace(0, 1) * 100).round(1)
                oh_risk_heatmap["OK %"] = (oh_risk_heatmap["OK"] / oh_risk_heatmap["Total"].replace(0, 1) * 100).round(1)

                fig_oh_heat = go.Figure()
                fig_oh_heat.add_trace(go.Bar(
                    y=oh_risk_heatmap["Proj_Lot"], x=oh_risk_heatmap["Short"], name="Short",
                    marker_color="#ef4444", orientation="h",
                    text=oh_risk_heatmap["Short %"].map(lambda v: f"{v:.1f}%"),
                    textposition="inside", textfont=dict(color="white", size=13, family="DM Sans"),
                ))
                fig_oh_heat.add_trace(go.Bar(
                    y=oh_risk_heatmap["Proj_Lot"], x=oh_risk_heatmap["OK"], name="OK",
                    marker_color="#10b981", orientation="h",
                    text=oh_risk_heatmap["OK %"].map(lambda v: f"{v:.1f}%"),
                    textposition="inside", textfont=dict(color="white", size=13, family="DM Sans"),
                ))
                fig_oh_heat.update_layout(barmode="stack", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8"), height=max(300, len(oh_risk_heatmap) * 35),
                    margin=dict(l=10, r=10, t=10, b=30), legend=dict(orientation="h", yanchor="bottom", y=-0.15))
                st.plotly_chart(fig_oh_heat, use_container_width=True)

                # Comparison summary: Full vs On-Hand Only
                st.markdown('<div class="section-header">📊 Full Allocation vs On-Hand Only Comparison</div>', unsafe_allow_html=True)
                comp_data = {
                    "Metric": ["Short Components", "OK Components", "Shortage %"],
                    "Full (OH+PO+Recv)": [short_comp, ok_comp, f"{shortage_pct:.1f}%"],
                    "On-Hand Only": [oh_short_comp, oh_ok_comp, f"{oh_shortage_pct:.1f}%"],
                    "Delta (OH Only − Full)": [
                        oh_short_comp - short_comp,
                        oh_ok_comp - ok_comp,
                        f"{oh_shortage_pct - shortage_pct:+.1f}%",
                    ],
                }
                st.dataframe(pd.DataFrame(comp_data), use_container_width=True, hide_index=True)

                st.download_button("⬇️ Download On-Hand Only Shortage Report (CSV)",
                    data=oh_risk_filtered.to_csv(index=False).encode("utf-8"),
                    file_name=f"onhand_only_shortage_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv",
                    key="dl_oh_shortage")

            st.markdown("---")
            st.download_button("⬇️ Download Priority Shortage Report (CSV)",
                data=filtered.to_csv(index=False).encode("utf-8"),
                file_name=f"priority_shortage_report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")
    else:
        st.warning("Required columns not found for shortage analysis.")


# ═══════════════════════════════════════════════════
# TAB 4 – LOT & PROJECT ANALYTICS
# ═══════════════════════════════════════════════════
with tab_analytics:
    st.markdown('<div class="section-header">Lot-wise and Project-wise Shortage Analytics</div>', unsafe_allow_html=True)
    st.caption("Shortage % = shortage unique component code count / total unique component code count, using sequential allocation from Shortage Forecast.")

    row_lot_assignments = load_row_lot_assignments()
    saved_lots = load_saved_lots()
    analytics_df = compute_priority_shortage_dataframe(df_fc, df_bom, saved_lots, row_lot_assignments)

    if analytics_df.empty:
        st.info("No analytics available yet. Assign lots in Forecasting tab to generate sequential shortage analytics.")
    else:
        analytics_df = analytics_df.copy()
        analytics_df["Is Shortage"] = analytics_df["Shortage"] > 0

        overall_total_unique = analytics_df["Component Code"].nunique()
        overall_short_unique = analytics_df[analytics_df["Is Shortage"]]["Component Code"].nunique()
        overall_short_pct = (overall_short_unique / overall_total_unique * 100) if overall_total_unique else 0
        lots_with_shortage = analytics_df[analytics_df["Is Shortage"]]["Lot"].nunique()

        km1, km2, km3, km4 = st.columns(4)
        with km1: render_metric("Lots Analyzed", analytics_df["Lot"].nunique(), "blue")
        with km2: render_metric("Projects Analyzed", analytics_df["Project"].nunique(), "cyan")
        with km3: render_metric("Lots With Shortages", lots_with_shortage, "red")
        with km4: render_metric("Overall Shortage %", f"{overall_short_pct:.1f}%", "purple")

        analytics_df["Assembly Start Date"] = pd.to_datetime(analytics_df["Assembly Start Date"], errors="coerce")
        date_source = analytics_df.dropna(subset=["Assembly Start Date"]).copy()

        if date_source.empty:
            st.info("Assembly Start Date is not available for date filtering.")
            date_filtered = analytics_df.copy()
        else:
            min_date = date_source["Assembly Start Date"].min().date()
            max_date = date_source["Assembly Start Date"].max().date()
            use_date_filter = st.checkbox("Apply Assembly Start Date filter", value=False, key="analytics_use_date_filter")
            if use_date_filter:
                selected_range = st.date_input(
                    "Filter by Assembly Start Date",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                    key="analytics_date_filter"
                )

                if isinstance(selected_range, tuple) and len(selected_range) == 2:
                    start_date, end_date = selected_range
                elif selected_range:
                    start_date = selected_range
                    end_date = selected_range
                else:
                    start_date = min_date
                    end_date = max_date

                start_ts = pd.Timestamp(start_date)
                end_ts = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)
                date_filtered = analytics_df[
                    (analytics_df["Assembly Start Date"] >= start_ts)
                    & (analytics_df["Assembly Start Date"] <= end_ts)
                ].copy()
            else:
                date_filtered = analytics_df.copy()

        project_options = sorted(date_filtered["Project"].dropna().astype(str).unique().tolist())
        lot_options = sorted(date_filtered["Lot"].dropna().astype(str).unique().tolist())

        af1, af2 = st.columns(2)
        with af1:
            analytics_proj_filter = st.multiselect(
                "Filter Projects (for tables + graphs)",
                project_options,
                default=[],
                key="analytics_proj_filter"
            )
        with af2:
            analytics_lot_filter = st.multiselect(
                "Filter Lots (for tables + graphs)",
                lot_options,
                default=[],
                key="analytics_lot_filter"
            )

        analytics_view = date_filtered.copy()
        if analytics_proj_filter:
            analytics_view = analytics_view[analytics_view["Project"].isin(analytics_proj_filter)]
        if analytics_lot_filter:
            analytics_view = analytics_view[analytics_view["Lot"].isin(analytics_lot_filter)]

        if analytics_view.empty:
            st.info("No records available for selected Project/Lot filters.")
        else:
            lot_rows = []
            for lot_nm, gdf in analytics_view.groupby("Lot"):
                total_unique = gdf["Component Code"].nunique()
                short_unique = gdf[gdf["Is Shortage"]]["Component Code"].nunique()
                pct = (short_unique / total_unique * 100) if total_unique else 0
                asm_dt = pd.to_datetime(gdf["Assembly Start Date"], errors="coerce").min()
                lot_rows.append({
                    "Assembly Start": asm_dt,
                    "Lot": lot_nm,
                    "Projects": gdf["Project"].nunique(),
                    "Total Unique Components": total_unique,
                    "Shortage Unique Components": short_unique,
                    "Shortage %": pct,
                })
            lot_summary = pd.DataFrame(lot_rows).sort_values(["Assembly Start", "Lot"], na_position="last")
            lot_summary_display = lot_summary.copy()
            lot_summary_display["Assembly Start"] = pd.to_datetime(lot_summary_display["Assembly Start"], errors="coerce").dt.strftime("%d-%b-%Y").fillna("")
            lot_summary_display["Shortage %"] = lot_summary_display["Shortage %"].map(lambda x: f"{x:.1f}%")

            st.markdown('<div class="section-header">Lot-wise Analytics (Assembly Start on Top)</div>', unsafe_allow_html=True)
            st.dataframe(lot_summary_display, use_container_width=True, height=320)

            proj_rows = []
            for proj_nm, gdf in analytics_view.groupby("Project"):
                total_unique = gdf["Component Code"].nunique()
                short_unique = gdf[gdf["Is Shortage"]]["Component Code"].nunique()
                pct = (short_unique / total_unique * 100) if total_unique else 0
                proj_rows.append({
                    "Project": proj_nm,
                    "Lots": gdf["Lot"].nunique(),
                    "Total Unique Components": total_unique,
                    "Shortage Unique Components": short_unique,
                    "Shortage %": pct,
                })
            proj_summary = pd.DataFrame(proj_rows).sort_values("Shortage %", ascending=False)
            proj_summary_display = proj_summary.copy()
            proj_summary_display["Shortage %"] = proj_summary_display["Shortage %"].map(lambda x: f"{x:.1f}%")

            st.markdown('<div class="section-header">Project-wise Analytics</div>', unsafe_allow_html=True)
            st.dataframe(proj_summary_display, use_container_width=True, height=320)

            st.markdown('<div class="section-header">Lot-wise Dual Analytics Graph</div>', unsafe_allow_html=True)
            lot_summary_plot = lot_summary.copy()
            lot_summary_plot["OK Unique Components"] = (
                lot_summary_plot["Total Unique Components"] - lot_summary_plot["Shortage Unique Components"]
            ).clip(lower=0)
            lot_summary_plot["OK %"] = (
                lot_summary_plot["OK Unique Components"] / lot_summary_plot["Total Unique Components"].replace(0, 1) * 100
            ).round(1)
            lot_summary_plot["Short %"] = lot_summary_plot["Shortage %"].round(1)

            fig_lot_dual = go.Figure()
            fig_lot_dual.add_trace(go.Bar(
                y=lot_summary_plot["Lot"],
                x=lot_summary_plot["OK Unique Components"],
                name="OK",
                marker_color="#10b981",
                orientation="h",
                text=lot_summary_plot["OK %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside",
                textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_lot_dual.add_trace(go.Bar(
                y=lot_summary_plot["Lot"],
                x=lot_summary_plot["Shortage Unique Components"],
                name="Short",
                marker_color="#ef4444",
                orientation="h",
                text=lot_summary_plot["Short %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside",
                textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_lot_dual.update_layout(
                barmode="stack",
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8"),
                height=max(320, len(lot_summary) * 28),
                margin=dict(l=10, r=10, t=10, b=30),
                legend=dict(orientation="h", yanchor="bottom", y=-0.2),
                xaxis_title="Component Count",
                yaxis_title=""
            )
            st.plotly_chart(fig_lot_dual, use_container_width=True)

            st.markdown('<div class="section-header">Project-wise Dual Analytics Graph</div>', unsafe_allow_html=True)
            proj_summary_plot = proj_summary.copy()
            proj_summary_plot["OK Unique Components"] = (
                proj_summary_plot["Total Unique Components"] - proj_summary_plot["Shortage Unique Components"]
            ).clip(lower=0)
            proj_summary_plot["OK %"] = (
                proj_summary_plot["OK Unique Components"] / proj_summary_plot["Total Unique Components"].replace(0, 1) * 100
            ).round(1)
            proj_summary_plot["Short %"] = proj_summary_plot["Shortage %"].round(1)

            fig_proj_dual = go.Figure()
            fig_proj_dual.add_trace(go.Bar(
                y=proj_summary_plot["Project"],
                x=proj_summary_plot["OK Unique Components"],
                name="OK",
                marker_color="#10b981",
                orientation="h",
                text=proj_summary_plot["OK %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside",
                textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_proj_dual.add_trace(go.Bar(
                y=proj_summary_plot["Project"],
                x=proj_summary_plot["Shortage Unique Components"],
                name="Short",
                marker_color="#ef4444",
                orientation="h",
                text=proj_summary_plot["Short %"].map(lambda v: f"{v:.1f}%"),
                textposition="inside",
                textfont=dict(color="white", size=13, family="DM Sans"),
            ))
            fig_proj_dual.update_layout(
                barmode="stack",
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8"),
                height=max(320, len(proj_summary) * 28),
                margin=dict(l=10, r=10, t=10, b=30),
                legend=dict(orientation="h", yanchor="bottom", y=-0.2),
                xaxis_title="Component Count",
                yaxis_title=""
            )
            st.plotly_chart(fig_proj_dual, use_container_width=True)


# ═══════════════════════════════════════════════════
# TAB 5 – TIMELINE
# ═══════════════════════════════════════════════════
with tab_timeline:
    if fc_proj_name_col:
        # ── Projects + Assigned Lots Schedule ──
        st.markdown('<div class="section-header">📦 Projects + Assigned Lots Schedule (by Assembly Start Date)</div>', unsafe_allow_html=True)
        st.caption("Only projects with assigned lots are shown, sorted by Assembly Start Date.")

        row_lot_assignments = load_row_lot_assignments()
        saved_lots = load_saved_lots()
        lot_schedule_rows = []
        for idx, row in df_fc.iterrows():
            rk = row.get("_row_key", "")
            assigned = [l for l in row_lot_assignments.get(rk, []) if l in saved_lots]
            if not assigned:
                continue
            pname = str(row.get(fc_proj_name_col, "Unknown")).strip() if pd.notna(row.get(fc_proj_name_col)) else "Unknown"
            asm_start = row.get(fc_asm_start_col) if fc_asm_start_col else None
            asm_end = row.get(fc_asm_end_col) if fc_asm_end_col else None
            need_by = row.get(fc_need_by_col) if fc_need_by_col else None
            bp = str(row.get(fc_build_period_col, "")).strip() if fc_build_period_col and pd.notna(row.get(fc_build_period_col)) else ""
            status = str(row.get(fc_status_col, "")).strip() if fc_status_col and pd.notna(row.get(fc_status_col)) else ""
            cab_qty = row.get(fc_cab_qty_col, 0) if fc_cab_qty_col else 0
            for lot_nm in assigned:
                lot_info = saved_lots.get(lot_nm, {})
                lot_metrics = compute_lot_metrics(df_bom, lot_info)
                lot_schedule_rows.append({
                    "Project": pname, "Build Period": bp, "Assembly Start": asm_start, "Assembly End": asm_end,
                    "Need By Date": need_by, "Status": status, "Cabinets": int(cab_qty) if pd.notna(cab_qty) else 0,
                    "Assigned Lot": lot_nm, "Lot Project": get_lot_project_name(lot_info) or "N/A",
                    "Lot ROW / ROM": get_lot_row_number(lot_info) or "N/A",
                    "Lot Components": lot_metrics["component_count"], "Lot Short": lot_metrics["short_components"],
                    "Lot Open Qty": lot_metrics["total_open_qty"],
                    "_asm_sort": asm_start if pd.notna(asm_start) else pd.Timestamp("2099-12-31"),
                })

        if lot_schedule_rows:
            lot_sched_df = pd.DataFrame(lot_schedule_rows).sort_values("_asm_sort").reset_index(drop=True)
            lot_sched_display = lot_sched_df.drop(columns=["_asm_sort"]).copy()
            for dcol in ["Assembly Start", "Assembly End", "Need By Date"]:
                if dcol in lot_sched_display.columns:
                    lot_sched_display[dcol] = pd.to_datetime(lot_sched_display[dcol], errors="coerce").dt.strftime("%d-%b-%Y").fillna("")

            lm1, lm2, lm3, lm4 = st.columns(4)
            with lm1: render_metric("Projects with Lots", lot_sched_display["Project"].nunique(), "blue")
            with lm2: render_metric("Total Lots Assigned", len(lot_sched_display), "purple")
            with lm3: render_metric("Total Short Components", int(lot_sched_display["Lot Short"].sum()), "red")
            with lm4: render_metric("Total Lot Open Qty", int(lot_sched_display["Lot Open Qty"].sum()), "amber")
            st.dataframe(lot_sched_display, use_container_width=True, height=400)

            gantt_rows = []
            for _, lrow in lot_sched_df.iterrows():
                asm_s = lrow["Assembly Start"]
                asm_e = lrow["Assembly End"]
                if pd.notna(asm_s):
                    end = asm_e if pd.notna(asm_e) else (asm_s + timedelta(days=7))
                    gantt_rows.append({"Label": f"{lrow['Project']} → {lrow['Assigned Lot']}", "Phase": "Assembly", "Start": asm_s, "End": end})
                nbd = lrow["Need By Date"]
                if pd.notna(nbd):
                    gantt_rows.append({"Label": f"{lrow['Project']} → {lrow['Assigned Lot']}", "Phase": "Shipment", "Start": nbd, "End": nbd + timedelta(days=3)})

            if gantt_rows:
                gantt_df = pd.DataFrame(gantt_rows)
                gantt_df["Start"] = pd.to_datetime(gantt_df["Start"], errors="coerce")
                gantt_df["End"] = pd.to_datetime(gantt_df["End"], errors="coerce")
                gantt_df = gantt_df.dropna(subset=["Start", "End"])

                fig_lot_tl = px.timeline(gantt_df, x_start="Start", x_end="End", y="Label", color="Phase",
                    color_discrete_map={"Assembly": "#3b82f6", "Shipment": "#f59e0b"})
                fig_lot_tl.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font=dict(color="#94a3b8"),
                    height=max(350, len(gantt_df["Label"].unique()) * 45), margin=dict(l=10, r=10, t=10, b=30),
                    legend=dict(orientation="h", yanchor="bottom", y=-0.15), xaxis_title="", yaxis_title="")
                fig_lot_tl.update_yaxes(autorange="reversed")

                st.plotly_chart(fig_lot_tl, use_container_width=True)
        else:
            st.info("No lots assigned to any forecasting rows yet. Assign lots in the Forecasting tab.")
    else:
        st.warning("Project Name column not found.")